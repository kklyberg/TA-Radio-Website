"""
Motorola Shop → Excel updater
-----------------------------
- Reads your existing Motorola.xlsx
- Scrapes product cards (part number, title, price) from a Motorola shop search URL
- Updates existing rows (price + short-description)
- Adds new part numbers that are not already in the spreadsheet
"""

import re
import time
from pathlib import Path
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# --------------------------------------------------
# CONFIGURATION – change these
# --------------------------------------------------
EXCEL_FILE   = r"C:\Users\kklyb\OneDrive\Documents\TA Radio Website\Testing Sand box1.xlsx"  # your existing spreadsheet
OUTPUT_FILE  = r"C:\Users\kklyb\OneDrive\Documents\TA Radio Website\Motorola_updated.xlsx"
SHEET_NAME   = "sheet1"          # change if your sheet has a different name

# The search / category page you want to scrape
SHOP_URL = "https://shop.motorolasolutions.com/search/_/N-735920568+387156806?Nrpp=15&srsltid=AfmBOoqLT-t-ndMuzn-9pT_2fw9-OLmtfLgeV-kj_LPaGwYG5_kSPgUQ"

# How many result pages to follow (set higher if needed)
MAX_PAGES = 3

# --------------------------------------------------
# HELPERS
# --------------------------------------------------
def clean_price(text: str) -> float | None:
    if not text:
        return None
    m = re.search(r"[\d,]+\.?\d*", text.replace(",", ""))
    if m:
        try:
            return float(m.group())
        except ValueError:
            return None
    return None

def normalize_part(part: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", part.upper())

# --------------------------------------------------
# SCRAPE THE SHOP PAGE
# --------------------------------------------------
def scrape_shop(url: str, max_pages: int = 3) -> list[dict]:
    products = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.set_default_timeout(30000)

        print(f"Opening {url}")
        page.goto(url, wait_until="networkidle")

        # Dismiss cookie banner if present
        try:
            page.click("text=CLOSE", timeout=3000)
        except Exception:
            pass
        try:
            page.click("button:has-text('Accept')", timeout=2000)
        except Exception:
            pass

        for page_num in range(1, max_pages + 1):
            print(f"  Scraping page {page_num} …")
            time.sleep(1.5)

            # Wait for product cards to appear
            page.wait_for_selector("text=ADD TO CART", timeout=15000)

            # Extract all product cards on the current page
            cards = page.query_selector_all("div[class*='product'], article, li[class*='product']")
            if not cards:
                # Fallback: look for any element that contains a price and a part number pattern
                cards = page.query_selector_all("body")

            # More reliable extraction using text content of the whole results area
            content = page.inner_text("body")

            # Find blocks that look like:  PARTNUMBER \n Title \n $price
            # This regex works well for the current Motorola shop layout
            pattern = re.compile(
                r"([A-Z0-9]{5,15}[A-Z0-9]?)\s*\n\s*(.+?)\s*\n\s*(?:In Stock|Out of Stock)?\s*\n?\s*\$([\d,]+\.?\d*)",
                re.MULTILINE
            )

            for match in pattern.finditer(content):
                part = match.group(1).strip()
                title = match.group(2).strip()
                price = clean_price(match.group(3))

                # Filter out obvious non-products
                if len(part) < 5 or "ADD TO CART" in title or "Results" in title:
                    continue

                products.append({
                    "part": part,
                    "title": title,
                    "price": price
                })

            # Try to go to the next page
            try:
                next_btn = page.query_selector("a[aria-label='Next'], button:has-text('Next'), a:has-text('>')")
                if next_btn and next_btn.is_enabled():
                    next_btn.click()
                    page.wait_for_load_state("networkidle")
                else:
                    break
            except Exception:
                break

        browser.close()

    # Deduplicate by part number
    seen = set()
    unique = []
    for p in products:
        key = normalize_part(p["part"])
        if key not in seen:
            seen.add(key)
            unique.append(p)

    print(f"Found {len(unique)} unique products")
    return unique

# --------------------------------------------------
# UPDATE THE EXCEL FILE
# --------------------------------------------------
def update_excel(products: list[dict]):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # Build a lookup of existing part numbers → row index
    existing = {}
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row=row_idx, column=1).value
        if model:
            existing[normalize_part(str(model))] = row_idx

    updated = 0
    added = 0

    for prod in products:
        part = prod["part"]
        title = prod["title"]
        price = prod["price"]
        key = normalize_part(part)

        if key in existing:
            # Update existing row
            row = existing[key]
            if price is not None:
                ws.cell(row=row, column=4).value = price          # price column
            if title:
                # Only overwrite short-description if it is empty or very short
                current = ws.cell(row=row, column=11).value
                if not current or len(str(current)) < 15:
                    ws.cell(row=row, column=11).value = title
                # Also put a longer version in catalog-copy if empty
                current_cat = ws.cell(row=row, column=13).value
                if not current_cat or len(str(current_cat)) < 20:
                    ws.cell(row=row, column=13).value = title
            updated += 1
            print(f"  Updated: {part}  →  ${price}  |  {title[:50]}")
        else:
            # Add new row
            new_row = [None] * 17
            new_row[0]  = part                # model
            new_row[1]  = "Motorola"          # brand
            new_row[2]  = "accessory"         # type (you can refine later)
            new_row[3]  = price               # price
            new_row[10] = title               # short-description
            new_row[12] = title               # catalog-copy
            new_row[13] = part                # root-model
            ws.append(new_row)
            added += 1
            print(f"  Added:   {part}  →  ${price}  |  {title[:50]}")

    wb.save(OUTPUT_FILE)
    print(f"\nFinished.")
    print(f"  Updated existing rows : {updated}")
    print(f"  New rows added        : {added}")
    print(f"  Saved to              : {OUTPUT_FILE}")

# --------------------------------------------------
# MAIN
# --------------------------------------------------
if __name__ == "__main__":
    print("Scraping Motorola shop …")
    products = scrape_shop(SHOP_URL, max_pages=MAX_PAGES)

    if not products:
        print("No products found. The page layout may have changed.")
        print("Try opening the URL in a browser and checking the structure.")
    else:
        print("\nUpdating spreadsheet …")
        update_excel(products)
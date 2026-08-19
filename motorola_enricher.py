"""
Motorola Part Number Enricher
-----------------------------
Searches the web for missing type / short-description / catalog-copy
and updates the spreadsheet.

Requirements:
    pip install openpyxl requests beautifulsoup4 duckduckgo-search lxml
"""

import time
import re
import openpyxl
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
from duckduckgo_search import DDGS   # optional but recommended

# --------------------------------------------------
# CONFIGURATION
# --------------------------------------------------
INPUT_FILE  = "Motorola.xlsx"          # your file
OUTPUT_FILE = "Motorola_enriched.xlsx" # safe output
SHEET_NAME  = "sheet1"                 # main data sheet

# Only process rows that are still empty or generic
TARGET_COLUMNS = {
    "type": 3,                 # column C
    "short-description": 11,   # column K
    "catalog-copy": 13,        # column M
}

# How many rows to process in one run (set high for full run)
MAX_ROWS_TO_PROCESS = 50          # start small, then raise
DELAY_BETWEEN_SEARCHES = 2.0      # seconds – be polite
TIMEOUT = 12

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

# --------------------------------------------------
# HELPERS
# --------------------------------------------------
def clean_text(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text).strip()
    return text[:300]          # keep it reasonable length

def guess_type_from_title(title: str) -> str:
    """Simple keyword → type mapping"""
    t = title.lower()
    rules = [
        (r"remote speaker mic|rsm|speaker microphone", "speaker-microphone"),
        (r"microphone|mic\b", "microphone"),
        (r"earpiece|earbud|ear set|earset", "earpiece"),
        (r"headset|head set", "headset"),
        (r"surveillance|2-wire|1-wire|acoustic tube", "surveillance-kit"),
        (r"battery|li-ion|nimh|lithium", "battery"),
        (r"charger|charging", "charger"),
        (r"antenna", "antenna"),
        (r"belt clip|carry case|nylon case|leather case|holster", "carry-accessory"),
        (r"programming cable|data cable|usb cable", "programming-cable"),
        (r"power cable|ignition sense", "power-cable"),
        (r"duplexer", "duplexer"),
        (r"repeater", "repeater"),
        (r"external speaker", "external-speaker"),
        (r"trunnion|mounting kit|footswitch", "vehicle-accessory"),
    ]
    for pattern, category in rules:
        if re.search(pattern, t):
            return category
    return "accessory"

def search_duckduckgo(part_number: str) -> dict:
    """Search DuckDuckGo and return best title + snippet + url"""
    query = f'Motorola "{part_number}" accessory OR radio'
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=5))
        if not results:
            return {}
        # Prefer results from motorola.com or motorolasolutions.com
        for r in results:
            href = r.get("href", "").lower()
            if "motorola" in href:
                return {
                    "title": r.get("title", ""),
                    "snippet": r.get("body", ""),
                    "url": r.get("href", "")
                }
        # fallback to first result
        r = results[0]
        return {
            "title": r.get("title", ""),
            "snippet": r.get("body", ""),
            "url": r.get("href", "")
        }
    except Exception as e:
        print(f"  Search error for {part_number}: {e}")
        return {}

def scrape_page(url: str) -> str:
    """Try to pull a cleaner product description from the page"""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        if resp.status_code != 200:
            return ""
        soup = BeautifulSoup(resp.text, "lxml")
        # Common product description selectors
        for selector in [
            "div.product-description", "div#product-description",
            "div.description", "meta[name=description]",
            "h1", "title"
        ]:
            el = soup.select_one(selector)
            if el:
                if el.name == "meta":
                    return clean_text(el.get("content", ""))
                return clean_text(el.get_text())
        return ""
    except Exception:
        return ""

# --------------------------------------------------
# MAIN
# --------------------------------------------------
def main():
    wb = load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]

    processed = 0
    updated = 0

    for row_idx in range(2, ws.max_row + 1):
        if processed >= MAX_ROWS_TO_PROCESS:
            break

        model = ws.cell(row=row_idx, column=1).value
        if not model:
            continue

        # Skip if we already have a decent short-description
        current_short = ws.cell(row=row_idx, column=TARGET_COLUMNS["short-description"]).value
        if current_short and len(str(current_short)) > 25:
            continue

        model = str(model).strip()
        print(f"[{processed+1}] Searching: {model}")

        result = search_duckduckgo(model)
        if not result:
            print("  → No results")
            processed += 1
            time.sleep(DELAY_BETWEEN_SEARCHES)
            continue

        title = clean_text(result.get("title", ""))
        snippet = clean_text(result.get("snippet", ""))
        url = result.get("url", "")

        # Optionally scrape the actual page for a better description
        page_desc = ""
        if url and "motorola" in url.lower():
            page_desc = scrape_page(url)

        description = page_desc or snippet or title
        item_type = guess_type_from_title(title + " " + description)

        # Write back
        if not ws.cell(row=row_idx, column=TARGET_COLUMNS["type"]).value:
            ws.cell(row=row_idx, column=TARGET_COLUMNS["type"]).value = item_type

        ws.cell(row=row_idx, column=TARGET_COLUMNS["short-description"]).value = title[:120]
        ws.cell(row=row_idx, column=TARGET_COLUMNS["catalog-copy"]).value = description[:400]

        print(f"  → type: {item_type}")
        print(f"  → {title[:80]}...")

        updated += 1
        processed += 1
        time.sleep(DELAY_BETWEEN_SEARCHES)

    wb.save(OUTPUT_FILE)
    print(f"\nFinished. Processed {processed} rows, updated {updated}.")
    print(f"Saved to: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()